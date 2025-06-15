from django.http import JsonResponse
from django.shortcuts import get_object_or_404
from django.utils.timezone import now
from django.views.decorators.csrf import csrf_exempt
from core.models import AdA, NFCChip
from django.utils.timezone import localtime
from django.core.files.storage import default_storage

from channels.layers import get_channel_layer
from asgiref.sync import async_to_sync

import json
import openpyxl

try:
    import smbus
except ImportError:
    smbus = None

try:
    import psutil
except ImportError:
    psutil = None

@csrf_exempt
def checkin_checkout(request):
    if request.method == "POST":
        data = json.loads(request.body)
        nfc_id = data.get("nfc_id")

        try:
            # NFC-Chip suchen
            nfc_chip = NFCChip.objects.get(nfc_id=nfc_id)
            ada = nfc_chip.ada

            if not ada:
                # NFC-Karte existiert, aber ist keinem AdA zugewiesen
                message = f"NFC-Karte {nfc_id} ist nicht zugewiesen. Jetzt registrieren?"
                send_websocket_message("warning", message, nfc_id)
                return JsonResponse({"error": message, "message_type": "warning", "nfc_id": nfc_id}, status=200)

        except NFCChip.DoesNotExist:
            # NFC-Karte existiert nicht -> Registrierung ermöglichen
            message = f"NFC-Karte {nfc_id} nicht registriert."
            send_websocket_message("danger", message, nfc_id)
            return JsonResponse({"error": message, "message_type": "warning", "nfc_id": nfc_id}, status=200)

        # Toggle Check-in / Check-out
        if ada.is_checked_in:
            ada.is_checked_in = False
            ada.last_checkout_time = now()
            action = "Check-out"
            message_type = "info"
        else:
            ada.is_checked_in = True
            ada.last_checkin_time = now()
            action = "Check-in"
            message_type = "success"

        ada.save()

        # WebSocket-Benachrichtigung senden
        message = f"{ada.rank} {ada.last_name} {action}"
        send_websocket_message(message_type, message)

        return JsonResponse({"message": message, "message_type": message_type})

# Hilfsfunktion zur WebSocket-Benachrichtigung
def send_websocket_message(message_type, message, nfc_id=None):
    channel_layer = get_channel_layer()
    async_to_sync(channel_layer.group_send)(
        "checkin_checkout",
        {
            "type": "checkin_checkout_message",
            "message": message,
            "message_type": message_type,
            "nfc_id": nfc_id,
        },
    )

def all_users(request):
    users = AdA.objects.all().values(
        "id", "rank", "last_name", "first_name", "last_checkin_time", "last_checkout_time", "is_checked_in"
    )

    formatted_users = []
    for user in users:
        check_time = (
            localtime(user["last_checkin_time"]).strftime("%d.%m.%Y %H:%M") if user["is_checked_in"]
            else localtime(user["last_checkout_time"]).strftime("%d.%m.%Y %H:%M")
        ) if user["last_checkin_time"] or user["last_checkout_time"] else "-"

        formatted_users.append({
            "id": user["id"],
            "rank": user["rank"],
            "last_name": user["last_name"],
            "first_name": user["first_name"],
            "status": "Eingecheckt" if user["is_checked_in"] else "Ausgecheckt",
            "check_time": check_time,
        })

    return JsonResponse({"data": formatted_users}, safe=False)

@csrf_exempt
def manual_checkout(request, ada_id):
    ada = get_object_or_404(AdA, id=ada_id)

    if not ada.is_checked_in:
        return JsonResponse({"error": "AdA ist bereits ausgecheckt."}, status=400)

    ada.is_checked_in = False
    ada.last_checkout_time = now()
    ada.save()

    return JsonResponse({"message": f"{ada.rank} {ada.last_name} wurde ausgecheckt."}, status=200)

@csrf_exempt
def add_ada(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            new_ada = AdA.objects.create(
                rank=data["rank"],
                last_name=data["last_name"],
                first_name=data["first_name"]
            )
            return JsonResponse({"message": "AdA erfolgreich hinzugefügt!", "id": new_ada.id}, status=201)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)


@csrf_exempt
def add_nfc_chip(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            ada_id = data.get("ada_id")
            nfc_id = data.get("nfc_id")

            # Prüfen, ob NFC bereits existiert
            if NFCChip.objects.filter(nfc_id=nfc_id).exists():
                return JsonResponse({"error": "NFC-Chip ist bereits registriert."}, status=400)

            ada = AdA.objects.get(id=ada_id)
            nfc_chip = NFCChip.objects.create(nfc_id=nfc_id, ada=ada)

            return JsonResponse({"message": "NFC-Chip erfolgreich hinzugefügt!", "id": nfc_chip.id}, status=201)
        except AdA.DoesNotExist:
            return JsonResponse({"error": "AdA nicht gefunden."}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

@csrf_exempt
def register_nfc_card(request):
    if request.method == "POST":
        data = json.loads(request.body)
        nfc_id = data.get("nfc_id")
        ada_id = data.get("ada_id")

        try:
            ada = AdA.objects.get(id=ada_id)
            nfc_chip = NFCChip.objects.create(nfc_id=nfc_id, ada=ada)

            return JsonResponse({"message": f"NFC-Karte {nfc_id} wurde {ada.rank} {ada.last_name} zugewiesen."}, status=201)
        except AdA.DoesNotExist:
            return JsonResponse({"error": "AdA nicht gefunden."}, status=400)

@csrf_exempt
def delete_ada(request, ada_id):
    if request.method == "DELETE":
        try:
            ada = AdA.objects.get(id=ada_id)
            ada.delete()
            return JsonResponse({"message": "AdA erfolgreich gelöscht!"}, status=200)
        except AdA.DoesNotExist:
            return JsonResponse({"error": "AdA nicht gefunden."}, status=404)
        except Exception as e:
            return JsonResponse({"error": str(e)}, status=400)

def get_ada_list(request):
    """API liefert eine Liste aller AdAs ohne NFC-Karte"""
    adas_without_nfc = AdA.objects.filter(nfc_chip__isnull=True).values("id", "rank", "last_name", "first_name")
    return JsonResponse({"adas": list(adas_without_nfc)}, safe=False)

def battery_status(request):
    try:
        bus = smbus.SMBus(1)
        ADDR = 0x2d

        # Ladezustand auslesen
        data = bus.read_i2c_block_data(ADDR, 0x02, 1)
        state_byte = data[0]
        if state_byte & 0x40:
            charging_state = "fast_charging"
        elif state_byte & 0x80:
            charging_state = "charging"
        elif state_byte & 0x20:
            charging_state = "discharging"
        else:
            charging_state = "idle"

        # VBUS: Spannung, Strom, Leistung
        data = bus.read_i2c_block_data(ADDR, 0x10, 6)
        voltage_mV = data[0] | (data[1] << 8)
        current_mA = data[2] | (data[3] << 8)
        power_mW = data[4] | (data[5] << 8)

        # Batterie: Spannung, Strom, % usw.
        data = bus.read_i2c_block_data(ADDR, 0x20, 6)
        battery_voltage = data[0] | (data[1] << 8)
        battery_current = data[2] | (data[3] << 8)
        if battery_current > 0x7FFF:
            battery_current -= 0x10000
        battery_percent = data[4] | (data[5] << 8)
        
        # CPU Temperature        
        f = open("/sys/class/thermal/thermal_zone0/temp", "r")
        cpu_temp = round(int(f.readline ())/1000.0,0)
        
        # CPU Percentage
        cpu_percent = psutil.cpu_percent()
        

        return JsonResponse({
            "charging_state": charging_state,
            "voltage_mV": voltage_mV,
            "current_mA": current_mA,
            "power_mW": power_mW,
            "battery_voltage": battery_voltage,
            "battery_current": battery_current,
            "battery_percent": battery_percent,
            "cpu_temp_celsius": cpu_temp,
            "cpu_percent": cpu_percent
        })

    except Exception as e:
        return JsonResponse({"error": str(e)}, status=500)

@csrf_exempt
def import_excel(request):
    if request.method == "POST":
        try:
            uploaded_file = request.FILES.get("file")
            if not uploaded_file:
                return JsonResponse({"error": "Keine Datei hochgeladen."}, status=400)

            # Datei temporär speichern
            file_path = default_storage.save(f"tmp/{uploaded_file.name}", uploaded_file)

            # Excel-Datei laden
            wb = openpyxl.load_workbook(default_storage.path(file_path))
            sheet = wb.active

            imported = 0
            header = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                row_data = dict(zip(header, row))
                grad = row_data.get("Grad Kurzform")
                name = row_data.get("Name / Vorname")
                if grad and name:
                    last_name = name.split(",")[0] if "," in name else name
                    first_name = name.split(",")[1].strip() if "," in name else ""
                    # Überprüfen, ob AdA bereits existiert
                    if AdA.objects.filter(rank=grad.strip(), last_name=last_name.strip()).exists():
                        continue

                    AdA.objects.create(rank=grad.strip(), last_name=last_name.strip(), first_name=first_name.strip())
                    imported += 1
            # Temporäre Datei löschen
            default_storage.delete(file_path)
            if imported == 0:
                return JsonResponse({"message": "Keine neuen AdA importiert."}) 
            return JsonResponse({"message": f"{imported} AdA(s) erfolgreich importiert."})

        except Exception as e:
            # Fehlerbehandlung
            if 'file_path' in locals():
                default_storage.delete(file_path)
            return JsonResponse({"error": str(e)}, status=500)